#!/usr/bin/env python3
"""
Post-process a *_weights.xlsx produced by the pipeline: flag which rows are
single-frame misreads (without ever altering or substituting a value - a
flagged row's weight is simply left out, everything else stays exactly as
read), detect and discard a leading invalid segment before a genuine reset/
retare event, a data-derived dispense-start-corrected time column, and a
second sheet converting the kept weight to remaining syringe volume (mL,
assuming water density). Everything is done as live Excel formulas (not
baked-in Python values) referencing a small, labeled parameter block, so
the thresholds are visible and adjustable directly in the spreadsheet.

History of the filter shape (why it looks like this):
  v1 - carry-forward filter (each row compared to the previous ACCEPTED
       value). Fatal flaw: one bad reading slipping through as "accepted"
       (iddsi2-10ml-1 @ t=100.4s) permanently drags down every later
       correct-but-now-lower reading too - the filter locks on and throws
       away the rest of the series.
  v2 - local median instead: each row judged against the median of a
       window of raw readings AROUND it, not a running value, so one bad
       row can't cascade. Window=7 initially.
  v3 - window widened 7 -> 12: a 4-consecutive-frame misread burst
       (iddsi2-20ml-2 @ t=99.2-99.5s, raw "061" four frames running)
       became the local majority in a 7-window and passed the filter
       untouched. 12 keeps bursts like that a minority while still
       resolving real multi-second transitions, and barely changes the
       clean file's exclusion count (iddsi2-10ml-1: 29 -> 31 of 2104).
  v4 - added leading-segment / reset detection (this version): some
       recordings start with a sustained WRONG plateau before the real
       experiment (iddsi2-20ml-2 spends its first ~20s reading ~8.5g when
       the true start is a hard reset down to ~1g at t=21s that then rises
       normally; iddsi2-20ml-1 shows the same shape at t=9-10s, and
       separately-confirmed visually to be a genuine OCR bug - the display
       reads "0.00" the whole time but a bad digit-box split decodes it as
       ~8.4-8.9g). A per-row local-window filter can't touch this because
       the whole plateau is internally consistent (nothing in it deviates
       from ITS OWN neighborhood) - it needs a different rule: find the
       single largest sustained drop anywhere in the trend (comparing each
       row's local median to the local median LOOKBACK rows earlier) and,
       if it exceeds the reset threshold, treat every row before it as
       invalid. Checked this doesn't false-trigger: iddsi2-10ml-1 (no
       reset, monotonic from the start) has a max drop of 0.01g, nowhere
       near the threshold, vs. 7.4-7.5g on the two files that do have one.
       Threshold raised 2.0 -> 5.0g after finding a false positive of its
       own: once the v7 classify_digit fix cleaned up iddsi2-20ml-1's
       leading segment, a dense burst of leading-digit-dropout misreads
       elsewhere in the same file (e.g. "3.23" flickering to "0.03") still
       transiently pulled the local median down by 3.18g - close enough to
       the old 2.0g threshold to wrongly flag a reset. 5.0g keeps safe
       margin below the real ~7.4g resets and above this noise ceiling.

Sheet1 (original data) gains:
  E: local_median_g   - MEDIAN of a window of raw weight readings centered
     on this row (window size in the parameter block). Reference value
     only, never a replacement.
  F: trend_ref_g       - local_median_g from LOOKBACK rows earlier (or this
     row's own value if there aren't enough rows yet) - the "before" side
     of the reset check below.
  G: reset_drop_g       - F minus E (positive = the trend has dropped since
     LOOKBACK rows ago).
  H: excluded           - 1 if this row's raw weight is blank, deviates
     from local_median_g by more than the tolerance, or falls before a
     detected reset point; else 0.
  I: weight_kept_g      - the ORIGINAL weight if not excluded, otherwise
     #N/A (which both Excel and LibreOffice charts treat as a real gap -
     see the note above the formula for why this isn't "" instead). Never
     a substituted or carried-forward number.
  J: is_started         - 1 once the raw weight first reaches the start
     threshold AND (if a reset was detected) this row is at/after it.
  K: time_corrected_s   - timestamp_sec minus the detected dispense-start
     time (first timestamp where is_started flips to 1) - negative before
     the detected start, 0 at start.
  Parameter block: window size, exclusion tolerance, start threshold,
     syringe volume, density, reset lookback/threshold, plus computed
     max-drop / reset-detected / reset-time / start-time cells - all
     editable/inspectable directly in the spreadsheet.

New sheet "Hacim" (Volume): time_corrected_s vs remaining volume in the
syringe (mL) = syringe_volume - weight_kept_g / density - #N/A wherever
weight_kept_g is #N/A, so excluded points leave a gap rather than an
interpolated/fabricated point. Points-only scatter chart (same style as
add_charts.py's weight chart).

Usage (from inside nestle_syringie/):
    ../.venv/bin/python excel_postprocess.py <path-to-weights.xlsx> [...]
"""
import sys
sys.path.append('/home/bee/projeler/lcd_reader/.venv/lib/python3.14/site-packages')

from pathlib import Path
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Font

MARKER_COLOR = "D62728"

MEDIAN_WINDOW = 20  # rows on each side (41-row window total) - widened from 12:
# on iddsi2-20ml-1 (after the v7 classify_digit fix), a ~20%-of-local-
# samples burst of a second digit-position 0/8 misread (raw "387"/"307"
# roughly 40:68 over a 20s span) locally outvoted the true value in some
# 12-wide windows (e.g. t=57.2s), and a "312"->"12" leading-digit-dropout
# burst around t=118-120s similarly slipped through. 20 resolves both
# (verified) while barely changing the clean file's exclusion count
# (iddsi2-10ml-1: 30 -> 31 of 2104). Cost: on iddsi2-20ml-2's genuine
# reset transition (~t=21s), one additional edge point right at the start
# of the transition gets excluded (528 -> 553 of 1251 total) - the
# transition itself stays clearly visible from its many other points, so
# this is an acceptable trade for cleaning up the denser noise elsewhere.
EXCLUDE_TOLERANCE_G = 0.15
START_THRESHOLD_G = 0.05
DEFAULT_DENSITY = 1.0
RESET_LOOKBACK_ROWS = 20  # ~2s at 10 samples/s
RESET_THRESHOLD_G = 5.0  # see module docstring: real resets measured at 7.4-7.5g; noise-driven false
# positives (a dense burst of leading-digit-dropout misreads, e.g. "3.23"
# flickering to "0.03", can transiently pull the local median down enough
# to look like a sustained drop) measured up to 3.18g on iddsi2-20ml-1
# after the v7 classify_digit fix - 5.0 keeps safe margin on both sides.


def infer_nominal_volume_ml(path):
    """Nominal syringe size from the filename convention (..._10ml-N /
    _20mL-N / _60mL-N ...) - falls back to 10 if it can't be parsed."""
    name = Path(path).stem.lower()
    for size in (60, 20, 10):
        if f"{size}ml" in name:
            return float(size)
    return 10.0


def add_postprocess(path):
    path = Path(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    n_rows = ws.max_row  # includes header row 1
    first_data_row = 2

    nominal_volume = infer_nominal_volume_ml(path)

    header_font = Font(bold=True)

    ws["E1"] = "local_median_g"
    ws["F1"] = "trend_ref_g"
    ws["G1"] = "reset_drop_g"
    ws["H1"] = "excluded"
    ws["I1"] = "weight_kept_g"
    ws["J1"] = "is_started"
    ws["K1"] = "time_corrected_s"
    for cell in ("E1", "F1", "G1", "H1", "I1", "J1", "K1"):
        ws[cell].font = header_font

    ws["N1"] = "Parametreler"
    ws["N1"].font = header_font
    ws["N2"] = "Medyan penceresi (satir, her yonde)"
    ws["O2"] = MEDIAN_WINDOW
    ws["N3"] = "Disleme toleransi (g)"
    ws["O3"] = EXCLUDE_TOLERANCE_G
    ws["N4"] = "Baslangic esigi (g)"
    ws["O4"] = START_THRESHOLD_G
    ws["N5"] = "Siringa hacmi (mL)"
    ws["O5"] = nominal_volume
    ws["N6"] = "Yogunluk (g/mL)"
    ws["O6"] = DEFAULT_DENSITY
    ws["N7"] = "Reset bakis penceresi (satir)"
    ws["O7"] = RESET_LOOKBACK_ROWS
    ws["N8"] = "Reset esigi (g)"
    ws["O8"] = RESET_THRESHOLD_G
    ws["N10"] = "En buyuk dusus (g)"
    ws["O10"] = f"=MAX(G{first_data_row}:G{n_rows})"
    ws["N11"] = "Reset tespit edildi mi (1/0)"
    ws["O11"] = f"=IF(O10>$O$8,1,0)"
    ws["N12"] = "Reset zamani (s)"
    ws["O12"] = f"=IF(O11=1,INDEX(B{first_data_row}:B{n_rows},MATCH(O10,G{first_data_row}:G{n_rows},0)),-1000000)"
    ws["N14"] = "Baslangic zamani (s)"
    ws["O14"] = f"=INDEX(B{first_data_row}:B{n_rows},MATCH(1,J{first_data_row}:J{n_rows},0))"

    for r in range(first_data_row, n_rows + 1):
        lo = max(first_data_row, r - MEDIAN_WINDOW)
        hi = min(n_rows, r + MEDIAN_WINDOW)
        ws[f"E{r}"] = f"=MEDIAN(D{lo}:D{hi})"

        ref_row = r - RESET_LOOKBACK_ROWS
        ws[f"F{r}"] = f"=E{ref_row}" if ref_row >= first_data_row else f"=E{r}"
        ws[f"G{r}"] = f"=F{r}-E{r}"

        ws[f"H{r}"] = (
            f"=IF(OR(NOT(ISNUMBER(D{r})),ABS(D{r}-E{r})>$O$3,"
            f"AND($O$11=1,B{r}<$O$12)),1,0)"
        )
        # NA() rather than "" - a formula returning "" is still plotted as
        # y=0 by Excel/LibreOffice scatter charts (a well-known charting
        # gotcha), which showed up as stray points along the bottom of the
        # chart at every excluded row instead of a gap. #N/A is the
        # standard technique both engines actually skip when plotting.
        ws[f"I{r}"] = f"=IF(H{r}=1,NA(),D{r})"
        ws[f"J{r}"] = f"=IF(AND(ISNUMBER(D{r}),D{r}>=$O$4,OR($O$11=0,B{r}>=$O$12)),1,0)"
        ws[f"K{r}"] = f"=B{r}-$O$14"

    # Kept weight vs corrected time - separate from the existing raw weight
    # chart so both are visible side by side for comparison. #N/A
    # (excluded) rows leave a gap rather than a fabricated point.
    ws._charts = ws._charts[:1]  # keep only the original raw-weight chart
    chart = ScatterChart()
    chart.title = f"{path.stem} - kept weight vs corrected time"
    chart.x_axis.title = "time since dispense start (s)"
    chart.y_axis.title = "weight (g)"
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.width = 24
    chart.height = 10
    chart.style = 2
    x_ref = Reference(ws, min_col=11, min_row=first_data_row, max_row=n_rows)  # K: time_corrected_s
    y_ref = Reference(ws, min_col=9, min_row=1, max_row=n_rows)  # I: weight_kept_g
    series = Series(y_ref, x_ref, title_from_data=True)
    series.marker = Marker(symbol="circle", size=5)
    series.marker.graphicalProperties = GraphicalProperties(solidFill=MARKER_COLOR)
    series.marker.graphicalProperties.line.solidFill = MARKER_COLOR
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    ws.add_chart(chart, "Q2")

    # --- New sheet: remaining syringe volume vs corrected time ---
    if "Hacim" in wb.sheetnames:
        del wb["Hacim"]
    vs = wb.create_sheet("Hacim")
    vs["A1"] = "time_corrected_s"
    vs["B1"] = "remaining_volume_mL"
    vs["A1"].font = header_font
    vs["B1"].font = header_font
    for r in range(first_data_row, n_rows + 1):
        vs[f"A{r}"] = f"=Sheet1!K{r}"
        vs[f"B{r}"] = f"=IF(ISNA(Sheet1!I{r}),NA(),Sheet1!$O$5-(Sheet1!I{r}/Sheet1!$O$6))"

    vchart = ScatterChart()
    vchart.title = f"{path.stem} - remaining syringe volume vs time"
    vchart.x_axis.title = "time since dispense start (s)"
    vchart.y_axis.title = "remaining volume (mL)"
    vchart.x_axis.axPos = "b"
    vchart.y_axis.axPos = "l"
    vchart.x_axis.delete = False
    vchart.y_axis.delete = False
    vchart.width = 24
    vchart.height = 10
    vchart.style = 2
    vx_ref = Reference(vs, min_col=1, min_row=first_data_row, max_row=n_rows)
    vy_ref = Reference(vs, min_col=2, min_row=1, max_row=n_rows)
    vseries = Series(vy_ref, vx_ref, title_from_data=True)
    vseries.marker = Marker(symbol="circle", size=5)
    vseries.marker.graphicalProperties = GraphicalProperties(solidFill="2CA02C")
    vseries.marker.graphicalProperties.line.solidFill = "2CA02C"
    vseries.graphicalProperties.line.noFill = True
    vchart.series.append(vseries)
    vs.add_chart(vchart, "D2")

    wb.save(path)
    print(f"updated: {path}  (nominal syringe volume assumed {nominal_volume:.0f} mL from filename)")


def main():
    paths = sys.argv[1:]
    if not paths:
        print("usage: excel_postprocess.py <weights.xlsx> [...]")
        return
    for p in paths:
        add_postprocess(p)


if __name__ == "__main__":
    main()
