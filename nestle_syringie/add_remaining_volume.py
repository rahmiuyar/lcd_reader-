#!/usr/bin/env python3
"""
Copy every *_weights.xlsx from batch_output/ into deney_verileri/, then add a
"remaining_mL" column (liquid left in the syringe, by time) plus a scatter
chart of it vs time. batch_output/ stays as the untouched raw pipeline
output; deney_verileri/ is the analyzed/finished experiment data.

remaining_mL = initial_mL - dispensed_weight_g / density_g_per_mL
Clipped to [0, initial_mL]: remaining volume can't physically go negative or
exceed what the syringe started with, which also tames the handful of known
still-uncaught outlier spike frames (e.g. an isolated misread reading
hundreds of grams too high) without needing a separate statistical filter.

initial_mL is parsed from the video's output-folder name (from
extract_lcd_weights_batch.py's video_key(): "<parent-folder>_<stem>", e.g.
"lvl2_10mL-1") by looking for a 10mL/20mL/60mL marker in it. The root-level
recordings (lcd_reader_1/2/3) don't have an mL marker in their name but were
confirmed via md5sum to be the same recordings as 10mL-1/2/3.mp4, so they're
special-cased to 10mL below.

Density is a fixed assumption (water, 1.0 g/mL) confirmed with the user -
change DENSITY_G_PER_ML below if a different liquid is used.
"""
import sys
sys.path.append('/home/bee/projeler/lcd_reader/.venv/lib/python3.14/site-packages')

import glob
import re
import shutil
from pathlib import Path
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties

DENSITY_G_PER_ML = 1.0

# root-level recordings have no mL marker in their name - confirmed via
# md5sum to be the 10mL-1/2/3 recordings
NO_MARKER_INITIAL_ML = {
    'lcd_reader_1': 10, 'lcd_reader_2': 10, 'lcd_reader_3': 10,
}

MARKER_COLOR = "D62728"

OUTPUT_DIR = Path(__file__).parent / "deney_verileri"
OUTPUT_DIR.mkdir(exist_ok=True)


def initial_volume_ml(key):
    if key in NO_MARKER_INITIAL_ML:
        return NO_MARKER_INITIAL_ML[key]
    match = re.search(r'(\d+)\s*mL', key, re.IGNORECASE)
    return int(match.group(1)) if match else None


for src_path in sorted(glob.glob(str(Path(__file__).parent / "batch_output" / "*" / "*_weights.xlsx"))):
    stem = Path(src_path).parent.name
    initial_ml = initial_volume_ml(stem)
    if initial_ml is None:
        print(f"skip (unknown initial volume): {src_path}")
        continue

    path = OUTPUT_DIR / Path(src_path).name
    shutil.copy2(src_path, path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    n_rows = ws.max_row

    ws['E1'] = 'remaining_mL'
    for row in range(2, n_rows + 1):
        w = ws.cell(row=row, column=4).value  # D: weight
        if w is None:
            ws.cell(row=row, column=5).value = None
            continue
        remaining = initial_ml - w / DENSITY_G_PER_ML
        remaining = max(0.0, min(initial_ml, remaining))
        ws.cell(row=row, column=5).value = round(remaining, 3)

    # keep charts already on the sheet (weight vs time from add_charts.py);
    # only remove a remaining_mL chart from a previous run of this script
    ws._charts = [c for c in ws._charts if getattr(c, 'title', None) is None
                  or 'remaining mL' not in str(c.title)]

    chart = ScatterChart()
    chart.title = f"{stem} - remaining mL in syringe vs time"
    chart.x_axis.title = "time (s)"
    chart.y_axis.title = "remaining volume (mL)"
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.width = 24
    chart.height = 10
    chart.style = 2

    x_ref = Reference(ws, min_col=2, min_row=2, max_row=n_rows)  # B: timestamp_sec
    y_ref = Reference(ws, min_col=5, min_row=1, max_row=n_rows)  # E: remaining_mL (incl. header)
    series = Series(y_ref, x_ref, title_from_data=True)
    series.marker = Marker(symbol="circle", size=5)
    series.marker.graphicalProperties = GraphicalProperties(solidFill=MARKER_COLOR)
    series.marker.graphicalProperties.line.solidFill = MARKER_COLOR
    series.graphicalProperties.line.noFill = True

    chart.series.append(series)
    ws.add_chart(chart, "F24")
    wb.save(path)
    print(f"remaining_mL + chart added: {path}  (initial={initial_ml}mL)")
